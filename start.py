# from selenium import webdriver
# from selenium.webdriver.common.keys import Keys
# from selenium.webdriver import chrome
# import array as arr
import select
# from openpyxl import workbook
import openpyxl

#
# def mann(fun):
#     def wrapper(*args,**kwargs):
#         print("==============")
#         fun(*args,**kwargs)
#         print("===============")
#     return wrapper
#
# @mann
# def x():
#     print("i'm trying to decorate")
# x()
#
# class Phone:
#     def __init__(self,ram,hdd,pro):
#         self.ram = ram
#         self.hdd = hdd
#         self.pro = pro
#
#     def cam(self,pix):
#         print("cam pixles",pix)
#         h=0
#
# app = Phone("8gb","128gb","qudcore")
# mic = Phone("16gb","256gb","snapdragon")
# print(app.__dict__)
# print(mic.pro)
# Phone.cam(99,98)

# class Pedha:
#     def __init__(self,land):
#         self.land = land
#     def aasthi(self,land):
#         land = 23
#
# class Chinnodu(Pedha):
#     def __init__(self,land,ast):
#         super().__init__(land)
#         self.ast=ast
#     def asst(self,ast):
#         self.ast = "own house"
#         print(ast)
#     def asst(self,ast,land):
#         print("total",ast,land)
#
# class Chinadhi(Pedha):
#     def dow(self):
#         print("jhfdgdhfx")
#
# c = Chinadhi(78)
#
# print(c.dow())
#
# class Budda(Chinnodu):
#     def __init__(self,land,ast,intre):
#         self.intre =intre
#         super().__init__(land,ast)
#
# bb = Budda(23,"own apart","3%")
#
# print(bb.__dict__)
# ---------------------------------------------------------
# def err(nu):
#     bh ="hhd"
#     try:
#         print(bh/nu)
#     except ZeroDivisionError:
#         print("exception caught")
#     except TypeError:
#         print("its type error")
#     finally:
#         try:
#             h = nu/0
#         except ZeroDivisionError:
#             print("exception caught in finally")
#
# err(0)
#--------------------------------------------------------------------

# p = r'C:\Users\DELL\Desktop\testxl.xlsx'
# work = openpyxl.load_workbook(p)
#
# sheet = work.get_sheet_by_name("Sheet1")
#
# r = sheet.max_row
# c = sheet.max_column
# for i in range(1,r+1):
#     for j in range(1,c+1):
#         if sheet.cell(i,j).value != None:
#             print("in if",sheet.cell(i,j).value)
#         else:
#             print("in else",sheet.cell(i,j).value)
# print(sheet.cell(1,1).value)


# x = int(input("enter x:"))
# y = int(input("enter y:"))
# for i in range(1,x):
#     for j in range(1,i):
#         print(" ",end="")
#     for k in range(1,y-i):
#         print("x",end=" ")
#     print("")


def is_leap(year):
    leap = False
    # print(year % 4 == 0, year(year % % 400 == 0, year % 100 != 0)
    if (year % 4 == 0) and  (year % 400 == 0 or year % 100 != 0):
        leap = True

    return leap


year = int(input("enter:"))
print(is_leap(year))